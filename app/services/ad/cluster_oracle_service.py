"""WB 顶级搜索集群 oracle service

数据流：
  用户在 WB 后台 "顶级搜索集群" 页面点"下载" → preset-stat-{adv}-{nm}_{from}-{to}.xlsx
  → 用户通过 UI 上传 → parse_preset_stat_xlsx → upsert_oracle
  → 聚类 API GET /clusters?nm_id=X 优先查 oracle，命中直接返
  → 未命中降级到 DeepSeek AI 三步聚类

xlsx 结构：
  Sheet 1 "Топ поисковых кластеров" 6 行 × 14 列 — 簇汇总（summary 表）
  Sheet 2 "words_cluster" 355 行 × 2 列 — 关键词 → 簇映射（oracle 表）
"""

from __future__ import annotations

import io
import re
from datetime import datetime, date as date_cls, timezone
from typing import Dict, List, Optional, Tuple

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.utils.logger import logger


SUMMARY_SHEET = "Топ поисковых кластеров"
MAPPING_SHEET = "words_cluster"

SUMMARY_HEADERS = [
    "Кластер", "Статус и тип ставки", "Ставка CPM", "Средняя позиция",
    "Просмотры", "Клики", "CTR", "Добавления в корзину",
    "Заказов с этим товаром", "Заказанные товары",
    "Затраты", "CPM", "CPC", "Валюта",
]

FILENAME_RE = re.compile(
    r"preset-stat-(\d+)-(\d+)_(\d{4}_\d{1,2}_\d{1,2})-(\d{4}_\d{1,2}_\d{1,2})",
    re.IGNORECASE,
)


def parse_filename_meta(filename: str) -> Tuple[Optional[int], Optional[int], Optional[date_cls], Optional[date_cls]]:
    """从 preset-stat-35895165-876156520_2026_4_17-2026_4_23.xlsx 抽 advert_id/nm_id/date 范围"""
    if not filename:
        return None, None, None, None
    m = FILENAME_RE.search(filename)
    if not m:
        return None, None, None, None
    advert_id = int(m.group(1))
    nm_id = int(m.group(2))

    def _parse_date(s):
        try:
            return datetime.strptime(s, "%Y_%m_%d").date()
        except Exception:
            return None

    return advert_id, nm_id, _parse_date(m.group(3)), _parse_date(m.group(4))


def _to_float(v):
    try:
        if v is None or v == "":
            return None
        if isinstance(v, str):
            v = v.replace(",", ".").strip()
            if v in ("-", "—", ""):
                return None
        return float(v)
    except Exception:
        return None


def _to_int(v):
    f = _to_float(v)
    return int(f) if f is not None else 0


def parse_preset_stat_xlsx(file_bytes: bytes) -> Tuple[List[Dict], List[Dict]]:
    """解析 xlsx 内容，返回 (summary_rows, mapping_rows)

    summary_rows: 每行是一个簇的汇总 dict
    mapping_rows: 每行是 {cluster_name, keyword}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    if SUMMARY_SHEET not in wb.sheetnames or MAPPING_SHEET not in wb.sheetnames:
        raise ValueError(
            f"xlsx 缺失必要 sheet：需要 {SUMMARY_SHEET!r} + {MAPPING_SHEET!r}，"
            f"实际有 {wb.sheetnames}"
        )

    # Sheet 1 summary
    ws1 = wb[SUMMARY_SHEET]
    summary_rows: List[Dict] = []
    for i, row in enumerate(ws1.iter_rows(values_only=True)):
        if i == 0:
            # 校验表头至少包含 Кластер 和 Просмотры
            hdr = [str(c or "").strip() for c in row]
            if "Кластер" not in hdr or "Просмотры" not in hdr:
                raise ValueError(f"xlsx Sheet {SUMMARY_SHEET!r} 表头不匹配: {hdr}")
            continue
        if not row or not row[0]:
            continue
        cluster_name = str(row[0]).strip()
        status = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        summary_rows.append({
            "cluster_name":  cluster_name,
            "status_type":   status,
            "is_excluded":   1 if (status and "Исключ" in status) else 0,
            "bid_cpm":       _to_float(row[2]) if len(row) > 2 else None,
            "avg_pos":       _to_float(row[3]) if len(row) > 3 else None,
            "views":         _to_int(row[4])   if len(row) > 4 else 0,
            "clicks":        _to_int(row[5])   if len(row) > 5 else 0,
            "ctr":           _to_float(row[6]) if len(row) > 6 else None,
            "baskets":       _to_int(row[7])   if len(row) > 7 else 0,
            "orders":        _to_int(row[8])   if len(row) > 8 else 0,
            "ordered_items": _to_int(row[9])   if len(row) > 9 else 0,
            "spend":         _to_float(row[10]) or 0 if len(row) > 10 else 0,
            "cpm":           _to_float(row[11]) if len(row) > 11 else None,
            "cpc":           _to_float(row[12]) if len(row) > 12 else None,
            "currency":      str(row[13]).strip() if len(row) > 13 and row[13] else "RUB",
        })

    # Sheet 2 mapping
    ws2 = wb[MAPPING_SHEET]
    mapping_rows: List[Dict] = []
    seen = set()
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 2 or not row[0] or not row[1]:
            continue
        cluster = str(row[0]).strip()
        kw      = str(row[1]).strip()
        if not cluster or not kw:
            continue
        key = (cluster, kw.lower())
        if key in seen:
            continue
        seen.add(key)
        mapping_rows.append({"cluster_name": cluster, "keyword": kw})

    return summary_rows, mapping_rows


def upsert_oracle(
    db: Session,
    tenant_id: int,
    shop_id: int,
    advert_id: int,
    nm_id: int,
    summary_rows: List[Dict],
    mapping_rows: List[Dict],
    date_from: Optional[date_cls],
    date_to: Optional[date_cls],
    source_file: Optional[str],
) -> Dict[str, int]:
    """把一批 summary + mapping 写入 oracle 表

    策略：
      - mapping：先 DELETE 本 (tenant,advert,nm) 旧数据 再批量 INSERT（全量替换，避免老数据漂移）
      - summary：按 UNIQUE (tenant,advert,nm,cluster,date_from,date_to) upsert
    """
    imported_at = datetime.now(timezone.utc)

    # 1. 全量替换 mapping
    db.execute(text("""
        DELETE FROM wb_cluster_oracle
        WHERE tenant_id=:tid AND advert_id=:adv AND nm_id=:nm
    """), {"tid": tenant_id, "adv": advert_id, "nm": nm_id})

    inserted_map = 0
    if mapping_rows:
        db.execute(text("""
            INSERT INTO wb_cluster_oracle
                (tenant_id, shop_id, advert_id, nm_id, cluster_name, keyword,
                 date_from, date_to, imported_at, source_file)
            VALUES
                (:tid, :sid, :adv, :nm, :cluster, :kw,
                 :df, :dt, :imp, :src)
        """), [
            {"tid": tenant_id, "sid": shop_id, "adv": advert_id, "nm": nm_id,
             "cluster": r["cluster_name"], "kw": r["keyword"],
             "df": date_from, "dt": date_to, "imp": imported_at, "src": source_file}
            for r in mapping_rows
        ])
        inserted_map = len(mapping_rows)

    # 2. upsert summary
    upserted_sum = 0
    for s in summary_rows:
        db.execute(text("""
            INSERT INTO wb_cluster_oracle_summary (
                tenant_id, shop_id, advert_id, nm_id, cluster_name,
                status_type, is_excluded, bid_cpm, avg_pos,
                views, clicks, ctr, baskets, orders, ordered_items,
                spend, cpm, cpc, currency, date_from, date_to,
                imported_at, source_file
            ) VALUES (
                :tid, :sid, :adv, :nm, :cluster,
                :status, :excl, :bid, :pos,
                :views, :clicks, :ctr, :baskets, :orders, :items,
                :spend, :cpm, :cpc, :cur, :df, :dt,
                :imp, :src
            )
            ON DUPLICATE KEY UPDATE
                tenant_id    = :tid,
                status_type  = :status,
                is_excluded  = :excl,
                bid_cpm      = :bid,
                avg_pos      = :pos,
                views        = :views,
                clicks       = :clicks,
                ctr          = :ctr,
                baskets      = :baskets,
                orders       = :orders,
                ordered_items= :items,
                spend        = :spend,
                cpm          = :cpm,
                cpc          = :cpc,
                currency     = :cur,
                imported_at  = :imp,
                source_file  = :src
        """), {
            "tid": tenant_id, "sid": shop_id, "adv": advert_id, "nm": nm_id,
            "cluster": s["cluster_name"], "status": s.get("status_type"),
            "excl": s.get("is_excluded", 0),
            "bid":   s.get("bid_cpm"),   "pos":    s.get("avg_pos"),
            "views": s.get("views", 0),  "clicks": s.get("clicks", 0),
            "ctr":   s.get("ctr"),
            "baskets": s.get("baskets", 0), "orders": s.get("orders", 0),
            "items": s.get("ordered_items", 0),
            "spend": s.get("spend", 0),  "cpm":    s.get("cpm"),
            "cpc":   s.get("cpc"),       "cur":    s.get("currency", "RUB"),
            "df": date_from, "dt": date_to,
            "imp": imported_at, "src": source_file,
        })
        upserted_sum += 1

    db.commit()
    logger.info(
        f"[cluster_oracle] upsert tenant={tenant_id} shop={shop_id} "
        f"advert={advert_id} nm={nm_id} mapping={inserted_map} summary={upserted_sum}"
    )
    return {"mapping": inserted_map, "summary": upserted_sum}


def get_oracle_summary(
    db: Session, tenant_id: int, advert_id: int, nm_id: int,
) -> List[Dict]:
    """取该 (tenant, advert, nm) 最新一次上传的簇汇总（按 imported_at desc 取最新 date 批次）"""
    rows = db.execute(text("""
        SELECT cluster_name, status_type, is_excluded, bid_cpm, avg_pos,
               views, clicks, ctr, baskets, orders, ordered_items,
               spend, cpm, cpc, currency, date_from, date_to, imported_at
        FROM wb_cluster_oracle_summary
        WHERE tenant_id=:tid AND advert_id=:adv AND nm_id=:nm
        ORDER BY imported_at DESC, views DESC
    """), {"tid": tenant_id, "adv": advert_id, "nm": nm_id}).mappings().all()

    if not rows:
        return []

    # 只返最新 imported_at 那批
    latest_imp = rows[0]["imported_at"]
    return [dict(r) for r in rows if r["imported_at"] == latest_imp]


def get_oracle_kw_to_cluster(
    db: Session, tenant_id: int, advert_id: int, nm_id: int,
) -> Dict[str, str]:
    """取 (tenant, advert, nm) 的 keyword(lower) → cluster_name 映射"""
    rows = db.execute(text("""
        SELECT cluster_name, keyword
        FROM wb_cluster_oracle
        WHERE tenant_id=:tid AND advert_id=:adv AND nm_id=:nm
    """), {"tid": tenant_id, "adv": advert_id, "nm": nm_id}).fetchall()
    return {r.keyword.lower().strip(): r.cluster_name for r in rows}


def oracle_last_imported(
    db: Session, tenant_id: int, advert_id: int, nm_id: int,
) -> Optional[datetime]:
    r = db.execute(text("""
        SELECT MAX(imported_at) AS t
        FROM wb_cluster_oracle
        WHERE tenant_id=:tid AND advert_id=:adv AND nm_id=:nm
    """), {"tid": tenant_id, "adv": advert_id, "nm": nm_id}).fetchone()
    return r.t if r and r.t else None
